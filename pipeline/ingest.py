"""
pipeline/ingest.py

Ingests ALL Excel files from data/raw/ into PostgreSQL.
Finds columns by header name rather than position — works across all DSLIB files
regardless of column layout differences.
"""

import openpyxl
import pandas as pd
from sqlalchemy import text
import os
import sys

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import DATA_RAW
from pipeline.db import engine


# ── HELPER: find a column by its header name ──────────────────────────────────

def find_column(ws, header_name, header_row=4, max_col=40):
    """
    Search a worksheet row for a header that contains header_name (case-insensitive).
    Returns the 1-based column index, or None if not found.
    """
    for col in range(1, max_col + 1):
        cell_value = ws.cell(header_row, col).value
        if cell_value and header_name.lower() in str(cell_value).lower():
            return col
    return None


def find_column_exact(ws, header_name, header_row=4, max_col=40):
    """
    Exact match version — for headers that share similar names
    e.g. 'Earned Value (EV)' vs 'Planned Value' vs 'Actual Cost'.
    Returns the 1-based column index, or None if not found.
    """
    for col in range(1, max_col + 1):
        cell_value = ws.cell(header_row, col).value
        if cell_value and str(cell_value).strip().lower() == header_name.strip().lower():
            return col
    return None


# ── STEP 1: Extract project-level info ────────────────────────────────────────

def extract_project_info(wb, filename):
    """
    Extract top-level project facts from the Baseline Schedule sheet.
    Searches for column headers rather than assuming fixed positions.
    """
    info = {
        'project_name':         filename.replace('.xlsx', ''),
        'budget_at_completion': None,
        'total_periods':        None,       # filled later after periods are counted
        'train_test_flag':      'train',    # default — update for test projects manually
        'project_size':         'medium'    # default — updated below based on BAC
    }

    # Try Baseline Schedule sheet first
    sheet_name = None
    for candidate in ['Baseline Schedule', 'Baseline', 'Schedule']:
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break

    if not sheet_name:
        print(f"  WARNING: No baseline schedule sheet found in {filename}")
        return info

    ws = wb[sheet_name]

    # Find the Total Cost or BAC column by header name
    # Try multiple possible header names used across different DSLIB files
    bac_col = None
    for candidate_header in ['Total Cost', 'Budget At Completion', 'BAC', 'Cost']:
        bac_col = find_column(ws, candidate_header, header_row=2)
        if bac_col:
            break

    if bac_col:
        # Row 3 is always the top-level project summary row in DSLIB baseline sheets
        bac_value = ws.cell(3, bac_col).value
        if bac_value and isinstance(bac_value, (int, float)):
            info['budget_at_completion'] = round(float(bac_value), 2)
    else:
        print(f"  WARNING: Could not find BAC column in {filename} — will be NULL")

    # Assign project size based on BAC
    if info['budget_at_completion']:
        bac = info['budget_at_completion']
        if bac < 500_000:
            info['project_size'] = 'small'
        elif bac < 5_000_000:
            info['project_size'] = 'medium'
        else:
            info['project_size'] = 'large'

    return info


# ── STEP 2: Extract period-by-period EV data ──────────────────────────────────

def extract_periods(wb, filename):
    """
    Loop through all TP sheets and extract PV, EV, AC for each period.
    Finds the correct columns by header name on each sheet.
    Falls back to the Tracking Overview sheet if TP sheets are missing.
    """

    # ── Option A: Extract from individual TP sheets ──────────────────────────
    tp_sheets = sorted(
        [s for s in wb.sheetnames if s.startswith('TP') and s[2:].isdigit()],
        key=lambda s: int(s[2:])
    )

    if tp_sheets:
        return _extract_from_tp_sheets(wb, tp_sheets, filename)

    # ── Option B: Fall back to Tracking Overview sheet ───────────────────────
    if 'Tracking Overview' in wb.sheetnames:
        print(f"  INFO: No TP sheets found — using Tracking Overview sheet")
        return _extract_from_tracking_overview(wb, filename)

    print(f"  WARNING: No period data found in {filename}")
    return pd.DataFrame()


def _extract_from_tp_sheets(wb, tp_sheets, filename):
    """Extract PV, EV, AC from individual TP1, TP2... sheets."""
    records = []

    # Detect column positions from the first TP sheet
    first_ws = wb[tp_sheets[0]]

    # TP sheets have headers on row 4 in DSLIB files
    header_row = 4

    ev_col = find_column(first_ws, 'Earned Value',  header_row)
    pv_col = find_column(first_ws, 'Planned Value', header_row)
    ac_col = find_column(first_ws, 'Actual Cost',   header_row)

    # If exact search fails, try broader search
    if not ev_col:
        ev_col = find_column(first_ws, 'EV', header_row)
    if not pv_col:
        pv_col = find_column(first_ws, 'PV', header_row)
    if not ac_col:
        ac_col = find_column(first_ws, 'AC', header_row)

    if not all([ev_col, pv_col, ac_col]):
        print(f"  WARNING: Could not detect EV/PV/AC columns in {filename}")
        print(f"    EV col: {ev_col}, PV col: {pv_col}, AC col: {ac_col}")
        return pd.DataFrame()

    print(f"  Detected columns — EV:{ev_col}, PV:{pv_col}, AC:{ac_col}")

    for sheet_name in tp_sheets:
        ws = wb[sheet_name]
        period_num  = int(sheet_name[2:])
        #status_date = ws.cell(1, 3).value

        # Row 5 is always the project-level summary (activity ID = 0)
        ev = ws.cell(5, ev_col).value
        pv = ws.cell(5, pv_col).value
        ac = ws.cell(5, ac_col).value

        # Skip periods with completely missing data
        if ev is None and pv is None and ac is None:
            continue

        records.append({
            'period_number': period_num,
            #'status_date':   status_date,
            'planned_value': float(pv) if pv is not None else None,
            'earned_value':  float(ev) if ev is not None else None,
            'actual_cost':   float(ac) if ac is not None else None,
        })

    df = pd.DataFrame(records).sort_values('period_number').reset_index(drop=True)
    return df


def _extract_from_tracking_overview(wb, filename):
    """
    Fall back: extract PV, EV, AC from the Tracking Overview sheet.
    Used for files that store data in summary form rather than per-TP sheets.
    """
    ws = wb['Tracking Overview']
    header_row = 2

    pv_col = find_column(ws, 'Planned Value', header_row)
    ev_col = find_column(ws, 'Earned Value',  header_row)
    ac_col = find_column(ws, 'Actual Cost',   header_row)

    if not all([ev_col, pv_col, ac_col]):
        print(f"  WARNING: Could not find PV/EV/AC in Tracking Overview for {filename}")
        return pd.DataFrame()

    records = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        try:
            period_num = int(row[0])
        except (ValueError, TypeError):
            continue

        records.append({
            'period_number': period_num,
            #'status_date':   None,
            'planned_value': float(row[pv_col - 1]) if row[pv_col - 1] is not None else None,
            'earned_value':  float(row[ev_col - 1]) if row[ev_col - 1] is not None else None,
            'actual_cost':   float(row[ac_col - 1]) if row[ac_col - 1] is not None else None,
        })

    return pd.DataFrame(records).sort_values('period_number').reset_index(drop=True)


# ── STEP 3: Save to PostgreSQL ─────────────────────────────────────────────────

def save_project(project_info):
    """Insert one project row and return the generated project_id."""
    with engine.connect() as conn:
        result = conn.execute(
            text("""
                INSERT INTO projects
                    (project_name, budget_at_completion, total_periods,
                     train_test_flag, project_size)
                VALUES
                    (:project_name, :budget_at_completion, :total_periods,
                     :train_test_flag, :project_size)
                RETURNING project_id
            """),
            project_info
        )
        conn.commit()
        return result.fetchone()[0]


def save_periods(periods_df, project_id):
    """Insert all period rows for one project."""
    periods_df = periods_df.copy()
    periods_df['project_id'] = project_id


    periods_df.to_sql(
        name='periods',
        con=engine,
        if_exists='append',
        index=False
    )


# ── STEP 4: Quality check ──────────────────────────────────────────────────────

def quality_check(periods_df, project_name):
    """
    Basic data quality checks after extraction.
    Prints warnings but does not stop ingestion.
    """
    total = len(periods_df)
    missing_ev = periods_df['earned_value'].isna().sum()
    missing_pv = periods_df['planned_value'].isna().sum()
    missing_ac = periods_df['actual_cost'].isna().sum()

    completeness = round((1 - missing_ev / total) * 100, 1) if total > 0 else 0

    print(f"  Quality check: {total} periods | completeness: {completeness}%")

    if missing_ev > 0:
        print(f"    WARNING: {missing_ev} periods missing EV")
    if missing_pv > 0:
        print(f"    WARNING: {missing_pv} periods missing PV")
    if missing_ac > 0:
        print(f"    WARNING: {missing_ac} periods missing AC")
    if completeness < 80:
        print(f"    DATA QUALITY GATE FAILED: completeness below 80% — project may be excluded from training")

    return completeness


# ── STEP 5: Check for duplicates ──────────────────────────────────────────────

def already_ingested(project_name):
    """Check if this project has already been loaded to avoid duplicates."""
    with engine.connect() as conn:
        result = conn.execute(
            text("SELECT project_id FROM projects WHERE project_name = :name"),
            {"name": project_name}
        )
        return result.fetchone() is not None


# ── MAIN: Process one file ─────────────────────────────────────────────────────

def ingest_file(filepath):
    """Full ingestion pipeline for one Excel file."""
    filename = os.path.basename(filepath)
    project_name = filename.replace('.xlsx', '')

    print(f"\nProcessing: {filename}")

    # Skip if already ingested
    if already_ingested(project_name):
        print(f"  SKIPPED — already in database")
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Extract
    project_info = extract_project_info(wb, filename)
    periods_df   = extract_periods(wb, filename)

    if periods_df.empty:
        print(f"  SKIPPED — no period data extracted")
        return None

    # Quality check
    completeness = quality_check(periods_df, project_name)

    # Update total periods count
    project_info['total_periods'] = len(periods_df)

    # Save to database
    project_id = save_project(project_info)
    save_periods(periods_df, project_id)

    print(f"  Saved — project_id: {project_id} | BAC: {project_info['budget_at_completion']} | periods: {len(periods_df)} | size: {project_info['project_size']}")
    return project_id


# ── MAIN: Process all files ────────────────────────────────────────────────────

def ingest_all():
    """Loop through every Excel file in data/raw/ and ingest it."""
    raw_folder = DATA_RAW

    if not os.path.exists(raw_folder):
        print(f"ERROR: Folder not found: {raw_folder}")
        return

    files = sorted([f for f in os.listdir(raw_folder) if f.endswith('.xlsx')])

    if not files:
        print("No Excel files found in data/raw/")
        return

    print(f"Found {len(files)} Excel file(s) in data/raw/\n")
    print("=" * 60)

    success = 0
    skipped = 0
    failed  = 0

    for filename in files:
        filepath = os.path.join(raw_folder, filename)
        try:
            result = ingest_file(filepath)
            if result is None:
                skipped += 1
            else:
                success += 1
        except Exception as e:
            print(f"  ERROR: {e}")
            failed += 1

    print("\n" + "=" * 60)
    print(f"Ingestion complete.")
    print(f"  Successful : {success}")
    print(f"  Skipped    : {skipped}")
    print(f"  Failed     : {failed}")


if __name__ == "__main__":
    ingest_all()